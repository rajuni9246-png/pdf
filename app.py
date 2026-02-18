import io
import os
import tempfile
import zipfile
from datetime import datetime

from flask import Flask, jsonify, make_response, request

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pdf2docx import Converter

app = Flask(__name__)


def _cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp


@app.after_request
def add_cors_headers(resp):
    return _cors(resp)


@app.route("/api/<path:_>", methods=["OPTIONS"])
def preflight(_):
    return ("", 204)


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({
        "ok": True,
        "converter": "pdf2docx + pymupdf + pdfplumber",
        "build": "2026-02-18-excel-text-v2",
        "excel_mode": "Text_P + Table_P",
        "pymupdf_loaded": bool(fitz),
        "time": datetime.utcnow().isoformat() + "Z"
    })


def _fitz_page_count(doc):
    return getattr(doc, "page_count", getattr(doc, "pageCount", 0))


def _fitz_load_page(doc, index):
    return doc.load_page(index) if hasattr(doc, "load_page") else doc.loadPage(index)


def _fitz_get_pixmap(page, scale=2.0, alpha=False):
    matrix = fitz.Matrix(scale, scale)
    if hasattr(page, "get_pixmap"):
        return page.get_pixmap(matrix=matrix, alpha=alpha)
    return page.getPixmap(matrix=matrix, alpha=alpha)


def _fitz_pix_to_bytes(pix, fmt):
    if hasattr(pix, "tobytes"):
        return pix.tobytes(fmt)
    return pix.getImageData(fmt)


@app.route("/api/pdf-to-word", methods=["POST"])
def pdf_to_word():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    src_path = None
    dst_path = None
    try:
        pdf_bytes = f.read()
        if not pdf_bytes:
            return jsonify({"error": "Uploaded file is empty"}), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as src:
            src.write(pdf_bytes)
            src_path = src.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as dst:
            dst_path = dst.name

        conversion_error = None
        try:
            cv = Converter(src_path)
            cv.convert(dst_path)
            cv.close()
        except Exception as ex:
            conversion_error = str(ex)

        if conversion_error or (not os.path.exists(dst_path)) or os.path.getsize(dst_path) == 0:
            return jsonify({"error": "Layout conversion failed: {}".format(conversion_error or "unknown error")}), 500
        else:
            with open(dst_path, "rb") as out:
                data = out.read()

        filename = os.path.splitext(f.filename or "converted")[0] + ".docx"
        resp = make_response(data)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        resp.headers["Content-Disposition"] = 'attachment; filename="{}"'.format(filename)
        return resp
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        for p in [src_path, dst_path]:
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass


@app.route("/api/pdf-to-excel", methods=["POST"])
def pdf_to_excel():
    return jsonify({"error": "PDF to Excel module removed in this build."}), 410

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        pdf_bytes = f.read()
        if not pdf_bytes:
            return jsonify({"error": "Uploaded file is empty"}), 400

        wb = Workbook()
        wb.remove(wb.active)

        ws_readme = wb.create_sheet(title="README")
        ws_readme["A1"] = "Use Text_P* sheets for editable page text."
        ws_readme["A2"] = "Use Table_P*_T* sheets for extracted tables."
        ws_readme["A3"] = "If only NoEditableText is present, the PDF is likely scanned/image-based."
        ws_readme.column_dimensions["A"].width = 110

        # Reliable editable text sheets from PyMuPDF text extraction.
        editable_count = 0
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_idx in range(_fitz_page_count(doc)):
            page = _fitz_load_page(doc, page_idx)
            text = ""
            try:
                text = page.get_text("text") if hasattr(page, "get_text") else page.getText("text")
            except Exception:
                text = ""

            plain = (text or "").strip()
            if plain:
                ws = wb.create_sheet(title=("Text_P{}".format(page_idx + 1))[:31])
                editable_count += 1
                for r_idx, line in enumerate(plain.splitlines(), start=1):
                    ws.cell(row=r_idx, column=1, value=line)
                ws.column_dimensions["A"].width = 120
        doc.close()

        # Table sheets provide editable extracted tabular data.
        table_count = 0
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_idx, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                for t_idx, table in enumerate(tables, start=1):
                    table_count += 1
                    ws = wb.create_sheet(title=("Table_P{}_T{}".format(page_idx, t_idx))[:31])
                    max_cols = 0
                    for r_idx, row in enumerate(table, start=1):
                        if row is None:
                            continue
                        max_cols = max(max_cols, len(row))
                        for c_idx, cell in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=(cell or "").strip())
                    for c_idx in range(1, max_cols + 1):
                        ws.column_dimensions[get_column_letter(c_idx)].width = 22

        if editable_count == 0 and table_count == 0:
            ws = wb.create_sheet(title="NoEditableText")
            ws["A1"] = "No editable text detected. This PDF may be scanned/image-based."
            ws["A2"] = "Use OCR-enabled conversion for editable output."

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = os.path.splitext(f.filename or "converted")[0] + ".xlsx"
        resp = make_response(out.getvalue())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = 'attachment; filename="{}"'.format(filename)
        return resp
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500


@app.route("/api/pdf-to-image", methods=["POST"])
def pdf_to_image():
    if fitz is None:
        return jsonify({"error": "PDF to image unavailable on this host (PyMuPDF not installed)."}), 501

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        pdf_bytes = f.read()
        if not pdf_bytes:
            return jsonify({"error": "Uploaded file is empty"}), 400

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        out = io.BytesIO()
        with zipfile.ZipFile(out, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for i in range(_fitz_page_count(doc)):
                page = _fitz_load_page(doc, i)
                pix = _fitz_get_pixmap(page, scale=2.0, alpha=False)
                zf.writestr("page-{}.png".format(i + 1), _fitz_pix_to_bytes(pix, "png"))

        doc.close()
        out.seek(0)

        filename = os.path.splitext(f.filename or "converted")[0] + "_images.zip"
        resp = make_response(out.getvalue())
        resp.headers["Content-Type"] = "application/zip"
        resp.headers["Content-Disposition"] = 'attachment; filename="{}"'.format(filename)
        return resp
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500


@app.route("/api/pdf-shrink", methods=["POST"])
def pdf_shrink():

    if fitz is None:
        return jsonify({"error": "PDF shrink unavailable (PyMuPDF not installed)."}), 500

    f = request.files.get("file")

    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    try:

        original = f.read()

        if not original:
            return jsonify({"error": "Uploaded file is empty"}), 400

        # Open PDF from memory
        src = fitz.open(stream=original, filetype="pdf")

        # Compress
        output = io.BytesIO()

        src.save(
            output,
            garbage=4,     # remove unused objects
            deflate=True,  # compress streams
            clean=True,    # clean structure
            linear=True    # optimize for web
        )

        src.close()

        compressed_bytes = output.getvalue()

        # Automatically choose smaller version
        if len(compressed_bytes) < len(original):
            final_bytes = compressed_bytes
        else:
            final_bytes = original

        filename = os.path.splitext(f.filename or "compressed")[0] + "_shrunk.pdf"

        response = make_response(final_bytes)

        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = f'attachment; filename=\"{filename}\"'

        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/pdf-merge", methods=["POST"])
def pdf_merge():
    if fitz is None:
        return jsonify({"error": "PDF merge unavailable on this host (PyMuPDF not installed)."}), 501

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    pdf_files = [x for x in files if x and (x.filename or "").lower().endswith(".pdf")]
    if len(pdf_files) < 2:
        return jsonify({"error": "Upload at least 2 PDF files"}), 400

    merged = fitz.open()
    opened_docs = []
    try:
        for f in pdf_files:
            doc = fitz.open(stream=f.read(), filetype="pdf")
            opened_docs.append(doc)
            merged.insert_pdf(doc)

        out = io.BytesIO()
        merged.save(out, garbage=4, deflate=True, clean=True, linear=True)
        out.seek(0)

        resp = make_response(out.getvalue())
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = 'attachment; filename="merged.pdf"'
        return resp
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        for doc in opened_docs:
            try:
                doc.close()
            except Exception:
                pass
        try:
            merged.close()
        except Exception:
            pass


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)

